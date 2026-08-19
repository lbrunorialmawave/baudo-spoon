"""Defensive parser for Fantagazzetta "rose" Excel exports.

Format observed on real fixtures (variable number of sheets / divisions):

- One sheet per division/girone (e.g. ``Divisione A``, ``Divisione B``, …).
  The parser iterates **all** sheets; no hardcoded count.
- Within a sheet, teams appear in contiguous 3-column blocks:
  ``[team_name | "costo" | (empty)]``.
- Row 1: team names + "costo" headers.
- Subsequent rows: ``(player_name, cost:int, None)`` until a row whose
  name cell is ``"totale"`` (case-insensitive) holding the spent budget.
- No role column, no Fantacalcio ID — only name + cost.
- Names may carry trailing markers such as ``*`` (injury / new signing flag
  from Fantagazzetta); these are stripped as noise.
- Empty teams (header + totale=0 only) are valid and returned with
  ``players=[]``.

The output is pure dataclasses / frozen structures ready for matching and
``RosterContext``; nothing is written to the database.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, Sequence

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────

#: Column stride of each team block.
BLOCK_WIDTH = 3

#: Case-insensitive marker that ends a team's player list.
TOTAL_MARKER = "totale"

#: Trailing markers Fantagazzetta may append to a player name.
_TRAILING_MARKER_RE = re.compile(r"\s*\*+\s*$")


# ── Data structures (runtime-only) ───────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class ParsedPlayer:
    """Single player entry as found in the export (pre-matching)."""

    name_raw: str
    """Original string from the cell (may still contain whitespace)."""

    name_clean: str
    """Name after stripping trailing markers and normalising whitespace."""

    cost: int
    """Purchase cost in credits (non-negative integer)."""

    row_index: int
    """1-based Excel row (useful for error reporting)."""


@dataclass(frozen=True, slots=True)
class ParsedTeam:
    """One fantasy team inside a division."""

    name: str
    """Display name of the team (header cell)."""

    players: tuple[ParsedPlayer, ...]
    """Players ordered as they appear in the sheet."""

    total_spent: int
    """Value of the ``totale`` cell (budget spent)."""

    column_index: int
    """0-based starting column of this team block."""

    is_empty: bool = field(init=False)

    def __post_init__(self) -> None:
        object.__setattr__(self, "is_empty", len(self.players) == 0)


@dataclass(frozen=True, slots=True)
class ParsedDivision:
    """One sheet / girone / fantasy league."""

    sheet_name: str
    """Exact sheet title (e.g. ``Divisione B``)."""

    teams: tuple[ParsedTeam, ...]
    """Teams in left-to-right order of appearance."""


@dataclass(frozen=True, slots=True)
class ParsedWorkbook:
    """Complete parse result of a multi-division rose export."""

    divisions: tuple[ParsedDivision, ...]
    """All sheets that contained at least one recognisable team header."""

    source_filename: str | None = None
    """Original filename when available (for logging / UI)."""

    @property
    def team_count(self) -> int:
        return sum(len(d.teams) for d in self.divisions)

    @property
    def player_count(self) -> int:
        return sum(
            len(t.players) for d in self.divisions for t in d.teams
        )


# ── Public API ───────────────────────────────────────────────────────────────


def parse_workbook(
    path: str | Path,
    *,
    source_filename: str | None = None,
) -> ParsedWorkbook:
    """Parse a Fantagazzetta rose Excel file from disk.

    Parameters
    ----------
    path:
        Path to ``.xlsx`` (or ``.xlsm``) file.
    source_filename:
        Optional override for the reported source name (defaults to
        ``Path(path).name``).

    Returns
    -------
    ParsedWorkbook
        Fully parsed, immutable structure. Never raises for empty teams
        or empty sheets; only raises on unreadable / unsupported formats.
    """
    path = Path(path)
    filename = source_filename or path.name
    log.info("Parsing rose workbook: %s", filename)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface as actionable error
        raise ValueError(
            f"Impossibile aprire il file Excel '{filename}': {exc}"
        ) from exc

    try:
        return _parse_workbook_obj(wb, source_filename=filename)
    finally:
        wb.close()


def parse_bytes(
    data: bytes | BinaryIO,
    *,
    source_filename: str | None = None,
) -> ParsedWorkbook:
    """Parse a rose Excel file from an in-memory buffer (upload path)."""
    if isinstance(data, bytes):
        stream: BinaryIO = io.BytesIO(data)
    else:
        stream = data

    filename = source_filename or "<upload>"
    log.info("Parsing rose workbook from bytes: %s", filename)

    try:
        wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            f"Impossibile aprire il file Excel '{filename}': {exc}"
        ) from exc

    try:
        return _parse_workbook_obj(wb, source_filename=filename)
    finally:
        wb.close()


# ── Internal implementation ──────────────────────────────────────────────────


def _parse_workbook_obj(
    wb: openpyxl.Workbook,
    *,
    source_filename: str | None,
) -> ParsedWorkbook:
    divisions: list[ParsedDivision] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        teams = _parse_sheet(ws, sheet_name=sheet_name)
        if teams:  # skip completely empty / non-roster sheets
            divisions.append(
                ParsedDivision(sheet_name=sheet_name, teams=tuple(teams))
            )
            log.info(
                "Sheet '%s': %d teams, %d players",
                sheet_name,
                len(teams),
                sum(len(t.players) for t in teams),
            )
        else:
            log.debug("Sheet '%s' produced no teams — skipped", sheet_name)

    if not divisions:
        raise ValueError(
            "Nessuna divisione/squadra riconosciuta nel file. "
            "Verifica che sia un export 'rose' di Fantagazzetta "
            "(blocchi di 3 colonne per squadra)."
        )

    result = ParsedWorkbook(
        divisions=tuple(divisions),
        source_filename=source_filename,
    )
    log.info(
        "Parse complete: %d divisions, %d teams, %d players",
        len(result.divisions),
        result.team_count,
        result.player_count,
    )
    return result


def _parse_sheet(ws: Worksheet, *, sheet_name: str) -> list[ParsedTeam]:
    """Extract all team blocks from a single worksheet."""
    # Materialise rows once (read_only iterator is one-shot).
    # We only need the first ~40 rows typically; cap at 200 for safety.
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 200:
            break
        rows.append(row)

    if not rows:
        return []

    header = rows[0]
    if not header:
        return []

    # Discover team start columns: cells that look like team names
    # (non-empty, not the literal "costo").
    team_starts: list[int] = []
    for col_idx, cell in enumerate(header):
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        if text.lower() == "costo":
            continue
        # A team name cell is expected at positions 0, 3, 6, …
        # We still accept any non-"costo" non-empty cell and treat it as
        # a potential team start; later validation checks the block.
        team_starts.append(col_idx)

    if not team_starts:
        log.warning(
            "Sheet '%s': no team headers found in first row", sheet_name
        )
        return []

    teams: list[ParsedTeam] = []
    for start_col in team_starts:
        team = _parse_team_block(rows, start_col=start_col, sheet_name=sheet_name)
        if team is not None:
            teams.append(team)

    return teams


def _parse_team_block(
    rows: Sequence[tuple],
    *,
    start_col: int,
    sheet_name: str,
) -> ParsedTeam | None:
    """Parse one 3-column team block starting at ``start_col``."""
    header_row = rows[0]
    if start_col >= len(header_row):
        return None

    raw_name = header_row[start_col]
    if raw_name is None:
        return None
    team_name = str(raw_name).strip()
    if not team_name or team_name.lower() == "costo":
        return None

    # Optional sanity: next cell should be "costo" (or empty / missing).
    costo_cell = (
        header_row[start_col + 1]
        if start_col + 1 < len(header_row)
        else None
    )
    if costo_cell is not None:
        costo_text = str(costo_cell).strip().lower()
        if costo_text and costo_text != "costo":
            log.debug(
                "Sheet '%s' col %d: expected 'costo' header, got %r — "
                "still attempting parse",
                sheet_name,
                start_col,
                costo_cell,
            )

    players: list[ParsedPlayer] = []
    total_spent = 0
    found_total = False

    for row_idx, row in enumerate(rows[1:], start=2):  # 1-based Excel rows
        if start_col >= len(row):
            continue

        name_cell = row[start_col]
        cost_cell = (
            row[start_col + 1] if start_col + 1 < len(row) else None
        )

        if name_cell is None and cost_cell is None:
            continue

        name_str = str(name_cell).strip() if name_cell is not None else ""

        # End-of-list marker
        if name_str.lower() == TOTAL_MARKER:
            total_spent = _safe_int(cost_cell, default=0)
            found_total = True
            break

        if not name_str:
            # Empty name with a cost? Ignore (noise).
            continue

        cost = _safe_int(cost_cell, default=None)
        if cost is None:
            log.warning(
                "Sheet '%s' team '%s' row %d: non-numeric cost %r for "
                "player %r — skipping row",
                sheet_name,
                team_name,
                row_idx,
                cost_cell,
                name_str,
            )
            continue

        if cost < 0:
            log.warning(
                "Sheet '%s' team '%s' row %d: negative cost %d for %r — "
                "clamping to 0",
                sheet_name,
                team_name,
                row_idx,
                cost,
                name_str,
            )
            cost = 0

        name_clean = _clean_player_name(name_str)
        if not name_clean:
            continue

        players.append(
            ParsedPlayer(
                name_raw=name_str,
                name_clean=name_clean,
                cost=cost,
                row_index=row_idx,
            )
        )

    if not found_total and players:
        # Some exports may omit the totale row; compute sum as fallback.
        total_spent = sum(p.cost for p in players)
        log.debug(
            "Sheet '%s' team '%s': no 'totale' row found — using sum of "
            "player costs (%d)",
            sheet_name,
            team_name,
            total_spent,
        )

    return ParsedTeam(
        name=team_name,
        players=tuple(players),
        total_spent=total_spent,
        column_index=start_col,
    )


def _clean_player_name(raw: str) -> str:
    """Strip Fantagazzetta trailing markers and collapse whitespace."""
    cleaned = _TRAILING_MARKER_RE.sub("", raw)
    cleaned = " ".join(cleaned.split())
    return cleaned


def _safe_int(value: object, *, default: int | None) -> int | None:
    """Convert cell value to int; return ``default`` on failure."""
    if value is None:
        return default
    if isinstance(value, bool):  # bool is subclass of int — reject
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return default
    try:
        text = str(value).strip().replace(",", ".")
        if not text:
            return default
        f = float(text)
        if f.is_integer():
            return int(f)
        return default
    except (TypeError, ValueError):
        return default
