"""Unit tests for the Fantagazzetta rose parser.

Uses the real Lido di Ostia fixture (3 divisions) as golden reference.
Also covers synthetic edge cases: empty teams, missing totale, markers, etc.
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from ml.roster_import.parser import (
    ParsedDivision,
    ParsedPlayer,
    ParsedTeam,
    ParsedWorkbook,
    parse_bytes,
    parse_workbook,
)

FIXTURE = Path(__file__).parent / "fixture_lido_di_ostia.xlsx"


# ── Real fixture ─────────────────────────────────────────────────────────────


def test_parse_real_fixture_structure():
    result = parse_workbook(FIXTURE)

    assert isinstance(result, ParsedWorkbook)
    assert result.source_filename == "fixture_lido_di_ostia.xlsx"
    assert len(result.divisions) == 3

    names = [d.sheet_name for d in result.divisions]
    assert names == ["Divisione A", "Divisione B", "Divisione C"]


def test_divisione_a_is_empty():
    result = parse_workbook(FIXTURE)
    div_a = result.divisions[0]
    assert div_a.sheet_name == "Divisione A"
    assert len(div_a.teams) == 12  # 12 team headers even if empty
    for team in div_a.teams:
        assert team.is_empty
        assert team.players == ()
        assert team.total_spent == 0


def test_divisione_b_team_counts_and_totals():
    result = parse_workbook(FIXTURE)
    div_b = next(d for d in result.divisions if d.sheet_name == "Divisione B")

    # 12 teams observed in the real file
    assert len(div_b.teams) == 12

    # Spot-check a few known teams and their totals from the Excel content
    by_name = {t.name: t for t in div_b.teams}

    assert "F.Q.F.C" in by_name
    fq = by_name["F.Q.F.C"]
    assert not fq.is_empty
    assert fq.total_spent == 474
    assert any(p.name_clean == "De Gea" and p.cost == 34 for p in fq.players)
    assert any(p.name_clean == "Calhanoglu" and p.cost == 60 for p in fq.players)

    assert "S.S. MTDC" in by_name
    mtdc = by_name["S.S. MTDC"]
    assert mtdc.total_spent == 452
    assert any(p.name_clean == "Hojlund" and p.cost == 154 for p in mtdc.players)

    assert "Ritalcarico F.C." in by_name
    rital = by_name["Ritalcarico F.C."]
    assert rital.total_spent == 500
    assert any(p.name_clean == "McTominay" and p.cost == 105 for p in rital.players)


def test_divisione_c_player_with_marker_stripped():
    """Cuenca A. * appears in some exports; marker must be removed."""
    result = parse_workbook(FIXTURE)
    # In the provided fixture the * is on Divisione B "Cuenca A. *" under Ritalcarico
    div_b = next(d for d in result.divisions if d.sheet_name == "Divisione B")
    rital = next(t for t in div_b.teams if t.name == "Ritalcarico F.C.")
    names = [p.name_clean for p in rital.players]
    assert "Cuenca A." in names
    # raw may still contain the star
    raw_with_star = [p for p in rital.players if "*" in p.name_raw]
    assert len(raw_with_star) >= 1
    for p in raw_with_star:
        assert "*" not in p.name_clean


def test_player_count_reasonable():
    result = parse_workbook(FIXTURE)
    # Div A empty, B and C have ~12 teams × ~25 players ≈ 600
    assert result.player_count > 500
    assert result.team_count == 12 + 12 + 12  # A also has 12 empty teams


def test_parse_bytes_same_as_path():
    data = FIXTURE.read_bytes()
    from_bytes = parse_bytes(data, source_filename="upload.xlsx")
    from_path = parse_workbook(FIXTURE)

    assert len(from_bytes.divisions) == len(from_path.divisions)
    assert from_bytes.team_count == from_path.team_count
    assert from_bytes.player_count == from_path.player_count
    assert from_bytes.source_filename == "upload.xlsx"


# ── Synthetic edge cases ─────────────────────────────────────────────────────


def _make_workbook(sheets: dict[str, list[list]]) -> bytes:
    """Build an in-memory xlsx from {sheet_name: list-of-rows}."""
    wb = openpyxl.Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_single_division_only():
    data = _make_workbook(
        {
            "Divisione A": [
                ["Alpha FC", "costo", None, "Beta FC", "costo", None],
                ["Player One", 10, None, "Player Two", 20, None],
                ["Player Three", 5, None, "totale", 20, None],
                ["totale", 15, None],
            ]
        }
    )
    result = parse_bytes(data)
    assert len(result.divisions) == 1
    assert result.divisions[0].sheet_name == "Divisione A"
    assert len(result.divisions[0].teams) == 2

    alpha = result.divisions[0].teams[0]
    assert alpha.name == "Alpha FC"
    assert len(alpha.players) == 2
    assert alpha.total_spent == 15
    assert alpha.players[0].name_clean == "Player One"
    assert alpha.players[0].cost == 10


def test_two_divisions():
    data = _make_workbook(
        {
            "Divisione A": [
                ["A1", "costo", None],
                ["P1", 1, None],
                ["totale", 1, None],
            ],
            "Divisione B": [
                ["B1", "costo", None],
                ["P2", 2, None],
                ["totale", 2, None],
            ],
        }
    )
    result = parse_bytes(data)
    assert len(result.divisions) == 2
    assert [d.sheet_name for d in result.divisions] == [
        "Divisione A",
        "Divisione B",
    ]


def test_empty_team_in_mixed_sheet():
    data = _make_workbook(
        {
            "Girone": [
                ["Full Team", "costo", None, "Empty Team", "costo", None],
                ["Some Player", 30, None, "totale", 0, None],
                ["totale", 30, None],
            ]
        }
    )
    result = parse_bytes(data)
    teams = result.divisions[0].teams
    assert len(teams) == 2
    full, empty = teams
    assert full.name == "Full Team"
    assert len(full.players) == 1
    assert full.total_spent == 30
    assert empty.name == "Empty Team"
    assert empty.is_empty
    assert empty.total_spent == 0


def test_missing_totale_row_uses_sum():
    data = _make_workbook(
        {
            "Solo": [
                ["NoTotal FC", "costo", None],
                ["A", 10, None],
                ["B", 25, None],
                # no totale row
            ]
        }
    )
    result = parse_bytes(data)
    team = result.divisions[0].teams[0]
    assert len(team.players) == 2
    assert team.total_spent == 35  # fallback sum


def test_trailing_star_stripped():
    data = _make_workbook(
        {
            "X": [
                ["T", "costo", None],
                ["Cuenca A. *", 1, None],
                ["Martinez Jo.", 50, None],
                ["totale", 51, None],
            ]
        }
    )
    result = parse_bytes(data)
    players = result.divisions[0].teams[0].players
    assert players[0].name_raw == "Cuenca A. *"
    assert players[0].name_clean == "Cuenca A."
    assert players[1].name_clean == "Martinez Jo."


def test_non_numeric_cost_skipped_with_warning(caplog):
    data = _make_workbook(
        {
            "X": [
                ["T", "costo", None],
                ["Good", 10, None],
                ["Bad", "abc", None],
                ["AlsoGood", 5, None],
                ["totale", 15, None],
            ]
        }
    )
    with caplog.at_level("WARNING"):
        result = parse_bytes(data)
    team = result.divisions[0].teams[0]
    assert len(team.players) == 2
    assert [p.name_clean for p in team.players] == ["Good", "AlsoGood"]
    assert any("non-numeric cost" in r.message for r in caplog.records)


def test_completely_invalid_file_raises():
    with pytest.raises(ValueError, match="Impossibile aprire"):
        parse_bytes(b"this is not an excel file")


def test_no_teams_raises():
    data = _make_workbook(
        {
            "EmptySheet": [
                [None, None, None],
                [None, None, None],
            ]
        }
    )
    with pytest.raises(ValueError, match="Nessuna divisione"):
        parse_bytes(data)


def test_frozen_dataclasses():
    result = parse_workbook(FIXTURE)
    team = result.divisions[1].teams[0]
    with pytest.raises(Exception):  # FrozenInstanceError or AttributeError
        team.name = "Hacked"  # type: ignore[misc]
