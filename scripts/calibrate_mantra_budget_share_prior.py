#!/usr/bin/env python3
"""Offline calibration of MANTRA department budget-share prior from listino.

Usage:
  python scripts/calibrate_mantra_budget_share_prior.py \
      [--xlsx quotazioni/Quotazioni_Fantacalcio_Stagione_2026_27.xlsx]
"""
from __future__ import annotations
import argparse, sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required", file=sys.stderr); sys.exit(1)

MANTRA_DEPARTMENTS = {
    "POR": ("Por",), "DIF": ("Dc", "B", "Dd", "Ds"),
    "CEN": ("E", "M", "C"), "TRQ": ("T", "W"), "ATT": ("A", "Pc"),
}
MANTRA_DEFAULT_QUOTAS = {
    "Por": 3, "Dc": 3, "B": 2, "Dd": 2, "Ds": 1,
    "E": 1, "M": 2, "C": 5, "T": 1, "W": 1, "A": 2, "Pc": 2,
}
DEPTH_ORDER = {
    "Por": 0, "Dc": 1, "B": 1, "Dd": 1, "Ds": 1,
    "E": 2, "M": 2, "C": 3, "T": 4, "W": 4, "A": 5, "Pc": 5,
}
ROLE_TO_DEPT = {r: d for d, roles in MANTRA_DEPARTMENTS.items() for r in roles}

def _dept_quota(d): return sum(MANTRA_DEFAULT_QUOTAS.get(r, 0) for r in MANTRA_DEPARTMENTS[d])

def _parse_rm(rm):
    if not rm: return []
    norm = {"POR":"Por","DC":"Dc","DD":"Dd","DS":"Ds","PC":"Pc"}
    out = []
    for p in str(rm).replace(",", ";").split(";"):
        p = p.strip()
        r = norm.get(p, norm.get(p.upper(), p))
        if r in DEPTH_ORDER: out.append(r)
    return out

def _primary_dept(roles):
    if not roles: return None
    return ROLE_TO_DEPT.get(min(roles, key=lambda r: DEPTH_ORDER.get(r, 99)))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("quotazioni/Quotazioni_Fantacalcio_Stagione_2026_27.xlsx"))
    ap.add_argument("--cost-col", choices=("auto","qt_a_m","qt_a"), default="auto")
    args = ap.parse_args()
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb["Tutti"] if "Tutti" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[1]]
    col = {h: i for i, h in enumerate(headers)}
    players = []
    for row in rows[2:]:
        if not row or row[0] is None: continue
        roles = _parse_rm(row[col.get("RM")])
        if not roles: continue
        qa, qam = row[col.get("Qt.A")], row[col.get("Qt.A M")] if "Qt.A M" in col else None
        if args.cost_col == "qt_a": cost = qa
        elif args.cost_col == "qt_a_m": cost = qam if qam is not None else qa
        else: cost = qam if (qam is not None and int(qam) > 0) else qa
        if cost is None or int(cost) <= 0: continue
        dept = _primary_dept(roles)
        if not dept: continue
        players.append({"name": str(row[col.get("Nome", 0)]), "dept": dept, "cost": int(cost)})
    by = defaultdict(list)
    for p in players: by[p["dept"]].append(p)
    raw, details = {}, {}
    for d in MANTRA_DEPARTMENTS:
        n = _dept_quota(d)
        top = sorted(by.get(d, []), key=lambda x: -x["cost"])[:n]
        raw[d] = float(sum(p["cost"] for p in top))
        details[d] = [(p["name"], p["cost"]) for p in top]
    total = sum(raw.values())
    shares = {k: round(v / total, 4) for k, v in raw.items()}
    drift = round(1.0 - sum(shares.values()), 4)
    if abs(drift) >= 0.0001:
        largest = max(shares, key=shares.get)
        shares[largest] = round(shares[largest] + drift, 4)
    print(f"# Source: {args.xlsx.name}")
    print("MANTRA_LISTINO_BUDGET_SHARE_PRIOR: dict[str, float] = {")
    for d in MANTRA_DEPARTMENTS:
        print(f'    "{d}": {shares[d]:.4f},  # top-{_dept_quota(d)} cost sum={raw[d]:.0f}')
    print("}")
    print(f"# sum = {sum(shares.values()):.4f}")
    for d in MANTRA_DEPARTMENTS:
        print(f"\n{d}:")
        for name, cost in details[d]:
            print(f"  {name:25s}  {cost:3d}")

if __name__ == "__main__":
    main()
